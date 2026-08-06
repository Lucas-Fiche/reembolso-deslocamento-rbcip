#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Importador RBCIP — planilha Google (xlsx) -> SQL para Supabase.
Gera: preview (stdout) + arquivo seed_migracao.sql (com PII -> NÃO versionar).
"""
import openpyxl, datetime, re, uuid, sys

F = sys.argv[1] if len(sys.argv) > 1 else "/root/.claude/uploads/8e9f28eb-626a-532b-9e9f-779135210bf2/621c208d-reembolsoDeslocamento.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "/tmp/claude-0/-home-user-reembolso-deslocamento-rbcip/8e9f28eb-626a-532b-9e9f-779135210bf2/scratchpad/seed_migracao.sql"
NS = uuid.UUID("11111111-2222-3333-4444-555555555555")  # namespace fixo -> ids estáveis

wb = openpyxl.load_workbook(F, data_only=True)
ws = wb["Registros"]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r and r[0]]

# índices
P,STA,CIDH,CILAT,CILNG,CIFOTO,CODH,COLAT,COLNG,COFOTO,TEMPO,NOME,CPF,EMAIL,TEL,VEIC,PLACA=range(17)
IDAPLAN,IDAT,IDAQ,VOLTAPLAN,VOLTAT,VOLTAQ,DIST,PEDT,QPED,VPED,COMPPED,USOU,PB,PR,CUPOM,VEST,VREAL,VTOT,VALID,RG,ORGAO,RECIBO,CARONAS,CONSUMO,FOTOCONS,AUDIT=range(17,43)

def num(x):
    if x is None: return None
    if isinstance(x,(int,float)): return float(x)
    s=str(x).strip().replace("R$","").replace(" ","").replace(",",".")
    try: return float(s)
    except: return None

def dt_iso(x):
    # Horários da planilha são de São Paulo (UTC-3, sem horário de verão).
    # Emitimos com o fuso explícito para o timestamptz gravar o instante certo.
    TZ="-03:00"
    if x is None or (isinstance(x,str) and not x.strip()): return None
    if isinstance(x,(datetime.datetime,datetime.date)):
        return x.strftime("%Y-%m-%d %H:%M:%S")+TZ
    s=str(x).strip()
    for fmt in ("%d/%m/%Y %H:%M:%S","%d/%m/%Y %H:%M","%Y-%m-%d %H:%M:%S","%d/%m/%Y"):
        try: return datetime.datetime.strptime(s,fmt).strftime("%Y-%m-%d %H:%M:%S")+TZ
        except: pass
    return None

def parse_trechos(txt, fase):
    """'N. origem → destino | KM km | data | GPS: .. | Maps: link' (por linha)."""
    out=[]; warn=[]
    if not txt or not str(txt).strip(): return out,warn
    for ln in str(txt).split("\n"):
        ln=ln.strip()
        if not ln: continue
        m=re.match(r'^\s*(\d+)\.\s*(.*)$', ln)
        ordem = int(m.group(1)) if m else len(out)+1
        body = m.group(2) if m else ln
        parts=[p.strip() for p in body.split("|")]
        origem=destino=None; km=None; horario=None; lat=lng=None; maps=None
        # 1ª parte: origem → destino
        if parts:
            od=parts[0]
            if "→" in od:
                a,b=od.split("→",1); origem=a.strip(); destino=b.strip()
            else:
                destino=od.strip()
        for pp in parts[1:]:
            low=pp.lower()
            if pp.endswith("km") or " km" in low:
                km=num(pp.replace("km",""));
            elif low.startswith("gps"):
                g=pp.split(":",1)[1].strip() if ":" in pp else ""
                if g and g.upper()!="N/A" and "," in g:
                    try: lat,lng=[float(z) for z in g.split(",")[:2]]
                    except: pass
            elif low.startswith("maps"):
                maps=pp.split(":",1)[1].strip() if ":" in pp else pp
            elif re.search(r'\d{2}/\d{2}/\d{4}|\d{4}-\d{2}-\d{2}', pp):
                horario=dt_iso(pp)
        if km is None: warn.append(f"km não lido em trecho {fase}: {ln[:50]}")
        out.append(dict(fase=fase,ordem=ordem,origem=origem,destino=destino,km=km or 0,
                        horario=horario,lat=lat,lng=lng,maps_link=maps))
    return out,warn

def parse_pedagios(txt, textcorr):
    """'N. [FASE] R$ VALOR | link' (por linha). Recupera valor do texto (conserta gremlin)."""
    out=[]; warn=[]
    if not txt or not str(txt).strip(): return out,warn
    for ln in str(txt).split("\n"):
        ln=ln.strip()
        if not ln: continue
        m=re.match(r'^\s*(\d+)\.\s*(?:\[(\w+)\])?\s*R\$ ?([\d.,]+)\s*(?:\|\s*(\S+))?', ln)
        if not m:
            warn.append(f"pedágio não lido: {ln[:50]}"); continue
        ordem=int(m.group(1)); fase=(m.group(2) or "").lower() or None
        valor=num(m.group(3)); link=m.group(4)
        if fase not in ("ida","volta"): fase=None
        out.append(dict(fase=fase,ordem=ordem,valor=valor or 0,comprovante_link=link))
    return out,warn

# ---- processa ----
STATUS_OK={"ABERTO","COMPLETO","REVISÃO","CORRIGIDO","APROVADO","PAGO","REPROVADO"}
recs=[]; all_warn=[]; km_mismatch=[]
for r in rows:
    proto=str(r[P]).strip()
    rid=str(uuid.uuid5(NS,proto))
    status=str(r[STA]).strip() if r[STA] else ""
    if status not in STATUS_OK: all_warn.append(f"{proto}: status inesperado '{status}'")
    ida,w1=parse_trechos(r[IDAT],"ida"); volta,w2=parse_trechos(r[VOLTAT],"volta")
    peds,w3=parse_pedagios(r[PEDT], r[COMPPED])
    all_warn+=[f"{proto}: {w}" for w in (w1+w2+w3)]
    dist=num(r[DIST])
    km_sum=sum((p["km"] or 0) for p in ida+volta)
    if dist and km_sum and abs(km_sum-dist)>1.0:
        km_mismatch.append((proto,dist,km_sum))
    rec=dict(id=rid,protocolo=proto,status=status,
        checkin_em=dt_iso(r[CIDH]),checkin_lat=num(r[CILAT]),checkin_lng=num(r[CILNG]),checkin_foto=r[CIFOTO],
        checkout_em=dt_iso(r[CODH]),checkout_lat=num(r[COLAT]),checkout_lng=num(r[COLNG]),checkout_foto=r[COFOTO],
        nome=(str(r[NOME]).strip() if r[NOME] else None),cpf=(str(r[CPF]).strip() if r[CPF] else None),
        email=(str(r[EMAIL]).strip() if r[EMAIL] else None),telefone=(str(r[TEL]).strip() if r[TEL] else None),
        veiculo=r[VEIC],placa=r[PLACA],rg=(str(r[RG]).strip() if r[RG] else None),orgao=r[ORGAO],
        dist_total=dist or 0,consumo=num(r[CONSUMO]) or 10,usou_estimativa=(str(r[USOU]).strip().lower()=="sim"),
        preco_base=num(r[PB]) or 6.47,preco_real=num(r[PR]),cupom_foto=r[CUPOM],
        validacao=(str(r[VALID]).strip() if r[VALID] else None),recibo_link=r[RECIBO],
        caronas=(str(r[CARONAS]).strip() if r[CARONAS] else None),foto_consumo=r[FOTOCONS],
        val_total_planilha=num(r[VTOT]),
        auditoria_txt=(str(r[AUDIT]).strip() if r[AUDIT] else None),
        criado_em=(dt_iso(r[CIDH]) or dt_iso(r[CODH])),
        paradas=ida+volta, pedagios=peds)
    recs.append(rec)

# ---- preview ----
from collections import Counter
st=Counter(x["status"] for x in recs)
tot_par=sum(len(x["paradas"]) for x in recs)
tot_ped=sum(len(x["pedagios"]) for x in recs)
print("╔═══════════════ PREVIEW DA IMPORTAÇÃO ═══════════════╗")
print(f"Reembolsos a criar : {len(recs)}")
print(f"  por status       : {dict(st)}")
print(f"Paradas a criar    : {tot_par}")
print(f"Pedágios a criar   : {tot_ped}")
print(f"Avisos de parsing  : {len(all_warn)}")
print(f"km (soma trechos) ≠ Dist Total (>1km): {len(km_mismatch)}")
print("╚═════════════════════════════════════════════════════╝")
if all_warn:
    print("\n-- AVISOS (até 15) --")
    for w in all_warn[:15]: print("  •",w)
if km_mismatch:
    print("\n-- km divergente (até 10) [proto, dist_total, soma_trechos] --")
    for x in km_mismatch[:10]: print("  •",x)

# amostra 1 registro completo
ex=next(x for x in recs if x["status"]=="PAGO" and x["pedagios"] and len(x["paradas"])>=2)
print("\n-- EXEMPLO de 1 reembolso decomposto --")
print("  reembolso:",ex["protocolo"],"| nome:",ex["nome"],"| dist:",ex["dist_total"],"| total_planilha:",ex["val_total_planilha"])
for p in ex["paradas"]: print("   parada:",p["fase"],p["ordem"],p["origem"],"→",p["destino"],"|",p["km"],"km |",p["horario"])
for p in ex["pedagios"]: print("   pedágio:",p["fase"],p["ordem"],"R$",p["valor"])

# ---- gera SQL ----
def q(v):
    if v is None or v=="" : return "NULL"
    if isinstance(v,bool): return "true" if v else "false"
    if isinstance(v,(int,float)): return repr(v)
    return "'"+str(v).replace("'","''")+"'"

lines=["-- MIGRAÇÃO RBCIP planilha->Supabase (gerado). Contém dados pessoais.",
       "begin;"]
for x in recs:
    lines.append(f"insert into reembolsos (id,protocolo,status,checkin_em,checkin_lat,checkin_lng,checkin_foto,checkout_em,checkout_lat,checkout_lng,checkout_foto,nome,cpf,email,telefone,veiculo,placa,rg,orgao,dist_total,consumo,usou_estimativa,preco_base,preco_real,cupom_foto,validacao,recibo_link,caronas,foto_consumo,val_total_planilha,criado_em) values ("
        + ",".join(q(x[k]) for k in ["id","protocolo","status","checkin_em","checkin_lat","checkin_lng","checkin_foto","checkout_em","checkout_lat","checkout_lng","checkout_foto","nome","cpf","email","telefone","veiculo","placa","rg","orgao","dist_total","consumo","usou_estimativa","preco_base","preco_real","cupom_foto","validacao","recibo_link","caronas","foto_consumo","val_total_planilha","criado_em"])
        + ");")
    for p in x["paradas"]:
        lines.append(f"insert into paradas (reembolso_id,fase,ordem,origem,destino,km,horario,lat,lng,maps_link) values ("
            + ",".join([q(x['id']),q(p['fase']),q(p['ordem']),q(p['origem']),q(p['destino']),q(p['km']),q(p['horario']),q(p['lat']),q(p['lng']),q(p['maps_link'])]) + ");")
    for p in x["pedagios"]:
        lines.append(f"insert into pedagios (reembolso_id,fase,ordem,valor,comprovante_link) values ("
            + ",".join([q(x['id']),q(p['fase']),q(p['ordem']),q(p['valor']),q(p['comprovante_link'])]) + ");")
    if x["auditoria_txt"]:
        lines.append(f"insert into auditoria (reembolso_id,usuario,acao) values ({q(x['id'])},'migracao',{q(x['auditoria_txt'])});")
# preços gasolina
wp=wb["Preços Gasolina"]
for r in wp.iter_rows(min_row=2, values_only=True):
    if not r or r[2] is None: continue
    ini=dt_iso(r[0]); fim=dt_iso(r[1]); val=num(r[2]); reg=r[3]
    ini = ini.split(" ")[0] if ini else None
    fim = fim.split(" ")[0] if fim else None
    lines.append(f"insert into precos_gasolina (inicio,fim,valor,registrado_por) values ({q(ini)},{q(fim)},{q(val)},{q(reg)});")
lines.append("commit;")
open(OUT,"w").write("\n".join(lines))
print(f"\nSQL gerado em: {OUT}  ({len(lines)} linhas)")
